from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from docnexus.ai.table_engine.core.models import (  # noqa: E402
    CanonicalDocument,
    FieldSpec,
    FileAsset,
    StructuredRecord,
    TargetTableSpec,
    TemplateSpec,
)
from docnexus.ai.table_engine.writers import XlsxWriter  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402


class XlsxWriterTests(unittest.TestCase):
    def test_date_datetime_values_are_written_without_time_suffix(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            template_path = work / "template.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["国家/地区", "日期", "病例数"])
            wb.save(template_path)

            template_doc = CanonicalDocument(
                doc_id="template",
                file=FileAsset(
                    id="template",
                    path=str(template_path),
                    name="template.xlsx",
                    ext="xlsx",
                    role="template",
                    mime_type=None,
                    size=template_path.stat().st_size,
                ),
                doc_type="xlsx",
            )
            template_spec = TemplateSpec(
                template_doc_id="template",
                target_tables=[
                    TargetTableSpec(
                        "table-1",
                        "covid",
                        schema=[
                            FieldSpec("country", "国家/地区", "国家地区", "string", True),
                            FieldSpec("date", "日期", "日期", "date", True),
                            FieldSpec("cases", "病例数", "病例数", "number", False),
                        ],
                    )
                ],
            )
            records = [
                StructuredRecord(
                    "r1",
                    "table-1",
                    values={"国家/地区": "Albania", "日期": datetime(2020, 2, 25), "病例数": 1},
                )
            ]

            result = XlsxWriter().write(template_doc, template_spec, records)
            out = load_workbook(result.output_path, data_only=True)
            value = out.active["B2"].value

            self.assertEqual(value, "2020-02-25")


if __name__ == "__main__":
    unittest.main()
