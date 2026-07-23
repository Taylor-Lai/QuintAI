import logging
from io import BytesIO
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class TableFiller:
    """Excel 表格自动填写器"""

    @staticmethod
    def fill_template(
        template_bytes: bytes,
        extracted_data: Dict,
        field_mapping: Dict[str, str] | None = None,
    ) -> BytesIO:
        """
        根据提取的数据填充 Excel 模板
        """
        try:
            # 1. 加载 Excel 模板
            wb = load_workbook(BytesIO(template_bytes))
            ws = wb.active

            logger.debug("Worksheet: %s", ws.title)
            logger.debug("Dimensions: %s", ws.dimensions)

            # 2. 如果没有提供映射关系，自动匹配表头
            if not field_mapping:
                field_mapping = TableFiller._auto_match_fields(ws, extracted_data)

            logger.debug("Field mapping: %s", field_mapping)
            logger.debug("Extracted data: %s", extracted_data)

            # 3. 获取表头（第一行）
            headers = []
            for cell in ws[1]:  # 第一行
                headers.append(str(cell.value).strip() if cell.value else "")

            logger.debug("Excel headers: %s", headers)

            # 4. 填充数据（从第 2 行开始，只填充第一行数据）
            filled_count = 0
            row_idx = 2  # 只填充第 2 行

            for col_idx, header in enumerate(headers, 1):
                if not header:
                    continue

                col_letter = get_column_letter(col_idx)

                # 检查是否有映射
                if header in field_mapping:
                    extract_field = field_mapping[header]

                    # 检查是否有数据
                    if extract_field in extracted_data:
                        value = extracted_data[extract_field]

                        # 处理嵌套字典或复杂结构
                        if isinstance(value, dict):
                            # 将字典转换为格式化的字符串
                            import json
                            value = json.dumps(value, ensure_ascii=False, indent=2)
                        elif isinstance(value, list):
                            # 将列表转换为字符串
                            value = ', '.join(str(v) for v in value)
                        else:
                            # 确保是字符串或数字
                            value = str(value) if value is not None else ""

                        cell_ref = f"{col_letter}{row_idx}"
                        ws[cell_ref] = value
                        logger.debug("Wrote %s [%s] = %s", cell_ref, header, value)
                        filled_count += 1
                    else:
                        logger.warning("No data for field %s", extract_field)
                else:
                    logger.warning("Header %s was not matched", header)

            logger.info("Filled %s cell(s)", filled_count)

            # 5. 保存到 BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return output

        except Exception as e:

            logger.exception("Table fill failed")
            raise Exception(f"表格填充失败：{str(e)}")

    @staticmethod
    def _auto_match_fields(ws, extracted_data: Dict) -> Dict[str, str]:
        """
        自动匹配 Excel 表头与提取字段
        """
        try:
            # 获取第一行表头
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            mapping = {}

            logger.debug("Excel headers: %s", headers)
            logger.debug("Extraction fields: %s", list(extracted_data.keys()))

            # 策略 1: 完全匹配
            for header in headers:
                if header in extracted_data:
                    mapping[header] = header
                    logger.debug("Exact field match: %s", header)

            # 策略 2: 包含匹配
            for header in headers:
                if header in mapping:
                    continue
                for extract_field in extracted_data.keys():
                    # 检查是否包含关系
                    if header in extract_field or extract_field in header:
                        mapping[header] = extract_field
                        logger.debug("Partial field match: %s -> %s", header, extract_field)
                        break

            # 策略 3: 关键词匹配
            keyword_map = {
                "姓名": ["学生姓名", "姓名", "负责人"],
                "教师": ["指导教师", "指导老师", "导师", "教师"],
                "学号": ["学号", "学生学号", "编号", "ID"],
                "课程": ["课程名称", "课程", "实验课程", "科目"],
                "时间": ["实验时间", "时间", "日期", "签署日期"],
                "地点": ["实验地点", "地点", "地址"],
            }

            for header in headers:
                if header in mapping:
                    continue
                for keyword, fields in keyword_map.items():
                    if keyword in header:
                        for field in fields:
                            if field in extracted_data:
                                mapping[header] = field
                                logger.debug("Keyword field match: %s (%s) -> %s", header, keyword, field)
                                break
                        if header in mapping:
                            break

            logger.debug("Final field mapping: %s", mapping)
            return mapping

        except Exception as e:
            logger.warning("Field matching failed: %s", e)
            return {}

    @staticmethod
    def create_template_from_fields(
        fields: List[str], output_path: str | None = None
    ) -> BytesIO:
        """
        根据字段列表自动生成 Excel 模板
        每个字段占一列
        """
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "数据填写表"

            # 逐个字段写入不同列。
            logger.debug("Creating template with fields: %s", fields)

            for col_idx, field in enumerate(fields, 1):  # 从第 1 列开始
                col_letter = get_column_letter(col_idx)
                ws[f"{col_letter}1"] = field  # 第一行是表头
                # 加粗表头
                ws[f"{col_letter}1"].font = ws[f"{col_letter}1"].font.copy(bold=True)
                logger.debug("Wrote header %s1: %s", col_letter, field)

            # 预留 10 行数据行（从第 2 行到第 11 行）
            for row in range(2, 12):
                for col in range(1, len(fields) + 1):
                    col_letter = get_column_letter(col)
                    ws[f"{col_letter}{row}"] = ""

            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # 保存
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info("Template created with %s column(s)", len(fields))
            return output

        except Exception as e:

            logger.exception("Template creation failed")
            raise Exception(f"创建模板失败：{str(e)}")
